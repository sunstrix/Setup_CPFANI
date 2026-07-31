# -*- coding: utf-8 -*-
"""gui.py - V5.9.8 (Edicao CP Fani: Flameshot dinamico via GitHub latest release)"""
from tkinter import messagebox
import customtkinter as ctk
import threading
import json
import os
import sys
import shutil
import subprocess
import urllib.request
import urllib.error
import re
import time
import traceback
import hashlib
import unicodedata
import io
from datetime import datetime
from pathlib import Path

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

try:
    from PIL import Image
    HAS_PIL = True
except ImportError:
    HAS_PIL = False
    print("[AVISO] PIL nao encontrado. Logo nao sera exibido.", flush=True)


def show_windows_toast(title, message):
    """Exibe notificacao nativa do Windows"""
    title_escaped = str(title).replace('"', '`"').replace("'", "`'")
    message_escaped = str(message).replace('"', '`"').replace("'", "`'")

    ps_script = rf'''
[Windows.UI.Notifications.ToastNotificationManager, Windows.UI.Notifications, ContentType = WindowsRuntime] | Out-Null
[Windows.UI.Notifications.ToastNotification, Windows.UI.Notifications, ContentType = WindowsRuntime] | Out-Null
[Windows.Data.Xml.Dom.XmlDocument, Windows.Data.Xml.Dom, ContentType = WindowsRuntime] | Out-Null
$appId = '{{1AC14E77-02E7-4E5D-B744-2EB1AE5198B7}}\WindowsPowerShell\v1.0\powershell.exe'
$template = @"
<toast>
<visual>
<binding template="ToastText02">
<text id="1">{title_escaped}</text>
<text id="2">{message_escaped}</text>
</binding>
</visual>
</toast>
"@
$xml = New-Object Windows.Data.Xml.Dom.XmlDocument
$xml.LoadXml($template)
$toast = New-Object Windows.UI.Notifications.ToastNotification $xml
[Windows.UI.Notifications.ToastNotificationManager]::CreateToastNotifier($appId).Show($toast)
'''

    try:
        subprocess.Popen(
            ["powershell", "-NoProfile", "-WindowStyle", "Hidden", "-Command", ps_script],
            creationflags=0x08000000 if sys.platform == "win32" else 0,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL
        )
    except Exception as e:
        print(f"[AVISO] Falha ao exibir notificacao: {e}", flush=True)


try:
    import mod_config
    import mod_instalar
except ImportError as e:
    msg = (
        f"[ERRO CRITICO] Falha ao importar modulos: {e}\n"
        "Certifique-se de que mod_config.py e mod_instalar.py estao no mesmo diretorio."
    )
    print(msg, flush=True)
    try:
        import tkinter as tk
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror("Erro Critico", msg)
        root.destroy()
    except Exception:
        pass
    sys.exit(1)

ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")

SETTINGS_PATH = os.path.join(os.path.dirname(__file__), "settings.json")


def load_settings():
    """Carrega configuracoes do settings.json com fallback seguro"""
    default_settings = {
        "apps": {
            "choco": [
                "googlechrome",
                "anydesk",
                "flameshot",
                "vlc",
                "winrar",
                "ditto",
                "vcredist-all",
                "rustdesk"
            ]
        },
        "bloatware_remove": [
            "Microsoft.ZuneVideo",
            "Microsoft.WindowsFeedbackHub"
        ]
    }

    if not os.path.exists(SETTINGS_PATH):
        print("[INFO] settings.json nao encontrado. Usando configuracoes padrao.", flush=True)
        return default_settings

    try:
        with open(SETTINGS_PATH, "r", encoding="utf-8", errors="replace") as f:
            settings = json.load(f)
        print(f"[OK] Configuracoes carregadas de {SETTINGS_PATH}", flush=True)
        return settings
    except json.JSONDecodeError as e:
        print(f"[ERRO] settings.json corrompido: {e}. Usando padrao.", flush=True)
        return default_settings
    except Exception as e:
        print(f"[ERRO] Falha ao ler settings.json: {e}. Usando padrao.", flush=True)
        return default_settings


SETTINGS = load_settings()


def _get_file_sha256(file_path, chunk_size=65536):
    try:
        h = hashlib.sha256()
        with open(file_path, "rb") as f:
            while True:
                chunk = f.read(chunk_size)
                if not chunk:
                    break
                h.update(chunk)
        return h.hexdigest().upper()
    except Exception as e:
        print(f"[ERRO] Falha ao calcular SHA256 de {file_path}: {e}", flush=True)
        return None


def _verify_sha256(file_path, expected_sha256):
    if not expected_sha256:
        print("[AVISO] Hash SHA256 esperado nao informado.", flush=True)
        return False

    actual = _get_file_sha256(file_path)
    if not actual:
        return False

    if actual.upper() != expected_sha256.strip().upper():
        print(f"[ERRO] Hash SHA256 invalido para {file_path}. Esperado: {expected_sha256}, Obtido: {actual}", flush=True)
        return False

    print(f"[OK] Hash SHA256 validado: {file_path}", flush=True)
    return True


def _get_expected_flameshot_sha256(version_tag=""):
    env_hash = os.environ.get("CPFANI_FLAMESHOT_MSI_SHA256", "").strip().upper()
    if env_hash:
        return env_hash

    resources_dir = os.path.join(os.path.dirname(__file__), "resources")
    candidates = []

    if version_tag:
        clean_version = str(version_tag).lstrip("vV").strip()
        if clean_version:
            candidates.append(f"Flameshot-{clean_version}-win64.msi.sha256")

    candidates.extend([
        "flameshot_msi.sha256",
        "Flameshot-13.3.0-win64.msi.sha256"
    ])

    for candidate in candidates:
        sidecar = os.path.join(resources_dir, candidate)
        if os.path.exists(sidecar):
            try:
                with open(sidecar, "r", encoding="ascii", errors="ignore") as f:
                    content = f.read().strip().split()[0].upper()
                    if content:
                        return content
            except Exception as e:
                print(f"[AVISO] Falha ao ler sidecar de hash do Flameshot {sidecar}: {e}", flush=True)

    return ""


def _version_to_list(v_str):
    nums = [int(x) for x in re.findall(r"\d+", str(v_str))]
    while len(nums) < 3:
        nums.append(0)
    return nums[:3]


def _normalize_snapshot_text(content):
    """Normaliza texto de snapshot para facilitar parsing robusto"""
    if isinstance(content, bytes):
        content = content.decode("utf-8", errors="replace")

    content = content.replace("\r\n", "\n").replace("\r", "\n")
    content = content.replace("\u00a0", " ")

    replacements = {
        "\u2014": "-",
        "\u2013": "-",
        "\u2212": "-",
        "\u00ba": "o",
        "\u00aa": "a",
        "N\u00ba": "Numero",
        "n\u00ba": "numero",
        "PERIF\u00c9RICOS": "PERIFERICOS",
        "Perif\u00e9ricos": "Perifericos",
        "Usu\u00e1rio": "Usuario",
        "usu\u00e1rio": "usuario",
        "N\u00e3o": "Nao",
        "n\u00e3o": "nao",
        "Edi\u00e7\u00e3o": "Edicao",
        "edi\u00e7\u00e3o": "edicao",
        "S\u00e9rie": "Serie",
        "s\u00e9rie": "serie",
        "Descri\u00e7\u00e3o": "Descricao",
        "descri\u00e7\u00e3o": "descricao",
        "Configura\u00e7\u00f5es": "Configuracoes",
        "configura\u00e7\u00f5es": "configuracoes",
        "Seguran\u00e7a": "Seguranca",
        "seguran\u00e7a": "seguranca",
        "Invent\u00e1rio": "Inventario",
        "invent\u00e1rio": "inventario",
        "Ger\u00eancia": "Gerencia",
        "ger\u00eancia": "gerencia",
    }

    for old, new in replacements.items():
        content = content.replace(old, new)

    content = unicodedata.normalize("NFKD", content)
    content = content.encode("ascii", "ignore").decode("ascii")
    return content


def _get_google_drive_service_and_snapshot_files():
    """
    Retorna (service, files) ja autenticado, ou (None, []) em caso de falha.
    Centraliza a logica de autenticacao OAuth2 e a busca dos arquivos
    CPFANI_Hardware_Snapshot* na pasta do Google Drive.
    """
    try:
        from google.oauth2.credentials import Credentials
        from google_auth_oauthlib.flow import InstalledAppFlow
        from google.auth.transport.requests import Request
        from googleapiclient.discovery import build
        from googleapiclient.errors import HttpError
        import pickle

        credentials_path = os.path.join(os.path.dirname(__file__), "credentials", "oauth2_credentials.json")
        token_path = os.path.join(os.path.dirname(__file__), "credentials", "token.pickle")

        print(f"[INFO] Drive read: credenciais em {credentials_path}", flush=True)
        print(f"[INFO] Drive read: token em {token_path}", flush=True)

        if not os.path.exists(credentials_path):
            print("[AVISO] Credenciais OAuth2 nao encontradas. Nao e possivel ler snapshots do Drive.", flush=True)
            return None, []

        SCOPES = ["https://www.googleapis.com/auth/drive"]
        required_scopes = set(SCOPES)
        creds = None

        if os.path.exists(token_path):
            try:
                with open(token_path, "rb") as token:
                    creds = pickle.load(token)
            except Exception as e:
                print(f"[ERRO] Falha ao ler token.pickle: {e}. Token sera recriado.", flush=True)
                creds = None

        if creds is not None:
            current_scopes = set(getattr(creds, "scopes", []) or [])
            if current_scopes and not required_scopes.issubset(current_scopes):
                print("[AVISO] Token atual nao possui escopo completo de Drive. Reautenticando...", flush=True)
                creds = None

        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                try:
                    creds.refresh(Request())
                    print("[OK] Token OAuth2 renovado com sucesso.", flush=True)
                except Exception as e:
                    print(f"[ERRO] Falha ao renovar token OAuth2: {e}. Reautenticando...", flush=True)
                    creds = None

            if not creds or not creds.valid:
                flow = InstalledAppFlow.from_client_secrets_file(credentials_path, SCOPES)
                creds = flow.run_local_server(port=0)
                print("[OK] Autenticacao OAuth2 concluida.", flush=True)

            try:
                with open(token_path, "wb") as token:
                    pickle.dump(creds, token)
            except Exception as e:
                print(f"[ERRO] Falha ao salvar token.pickle: {e}", flush=True)

        service = build("drive", "v3", credentials=creds)
        FOLDER_ID = "1EldWrM7U2tP4SPoGczMJyNdIIIcCsX3d"

        try:
            folder = service.files().get(fileId=FOLDER_ID, fields="id, name, capabilities").execute()
            print(f"[OK] Pasta do Drive acessivel: {folder.get('name')}", flush=True)
        except HttpError as e:
            print(f"[ERRO] Erro ao acessar pasta do Drive {FOLDER_ID}: {e}", flush=True)
            print("[AVISO] Verifique se a conta autenticada tem permissao na pasta.", flush=True)
            return None, []

        query = f"name contains 'CPFANI_Hardware_Snapshot_' and '{FOLDER_ID}' in parents and trashed=false"
        results = service.files().list(q=query, fields="files(id, name)").execute()
        files = results.get("files", [])
        return service, files
    except ImportError as e:
        print(f"[ERRO] Bibliotecas do Google Drive nao instaladas ou falha ao importar: {e}", flush=True)
        traceback.print_exc()
        return None, []
    except HttpError as e:
        print(f"[ERRO] Erro na API do Google Drive: {e}", flush=True)
        traceback.print_exc()
        return None, []
    except Exception as e:
        print(f"[ERRO] Erro ao conectar ao Google Drive: {e}", flush=True)
        traceback.print_exc()
        return None, []



def get_google_drive_service_and_snapshot_files():
    """Compatibilidade com nome antigo da funcao"""
    return _get_google_drive_service_and_snapshot_files()


def _parse_monitors_from_hardware_snapshot(content):
    """Extrai dados de monitores do conteudo do snapshot de hardware."""
    monitors = []

    try:
        normalized = _normalize_snapshot_text(content)

        if "PERIFERICOS - MONITORES" not in normalized:
            return monitors

        start_idx = normalized.find("PERIFERICOS - MONITORES")
        monitor_section = normalized[start_idx:]

        monitor_pattern = re.compile(
            r"Monitor\s+(\d+)\s*:\s*\n"
            r"\s*Modelo\s*:\s*(.*?)\s*\n"
            r"\s*Numero[_\s]+de[_\s]+Serie\s*:\s*(.*?)\s*"
            r"(?=\n\s*Monitor\s+\d+\s*:|\n={10,}|$)",
            re.DOTALL | re.IGNORECASE
        )

        matches = monitor_pattern.findall(monitor_section)

        for match in matches:
            monitor_num = match[0]
            modelo = match[1].strip()
            serial = match[2].strip()

            if modelo or serial:
                monitors.append({
                    "Numero_Monitor": int(monitor_num),
                    "Modelo": modelo if modelo else "N/A",
                    "Serial": serial if serial else "N/A"
                })

        if not monitors:
            print("[AVISO] Nenhum monitor encontrado no snapshot apos parsing.", flush=True)
    except Exception as e:
        print(f"[AVISO] Erro ao parsear monitores do snapshot: {e}", flush=True)

    return monitors


def _read_monitors_from_hardware_snapshots():
    """Le arquivos CPFANI_Hardware_Snapshot*.txt do Google Drive e extrai dados de monitores."""
    monitors_data = []

    service, files = _get_google_drive_service_and_snapshot_files()
    if service is None:
        return monitors_data

    if not files:
        print("[AVISO] Nenhum arquivo de snapshot encontrado no Google Drive.", flush=True)
        show_windows_toast("Aviso no Snapshot", "Nenhum arquivo de snapshot encontrado no Google Drive.")
        return monitors_data

    for file in files:
        try:
            from googleapiclient.http import MediaIoBaseDownload

            request = service.files().get_media(fileId=file["id"])
            fh = io.BytesIO()
            downloader = MediaIoBaseDownload(fh, request)
            done = False

            while done is False:
                status, done = downloader.next_chunk()

            content = fh.getvalue().decode("utf-8", errors="replace")
            normalized = _normalize_snapshot_text(content)

            pc_name_match = re.search(r"Nome_Computador\s*:\s*(.*?)\s*\n", normalized)
            pc_name = pc_name_match.group(1) if pc_name_match else "UNKNOWN"

            date_match = re.search(r"Gerado em:\s*(.*?)\s*\n", normalized)
            snapshot_date = date_match.group(1) if date_match else "N/A"

            monitors = _parse_monitors_from_hardware_snapshot(normalized)

            if not monitors:
                print(f"[AVISO] Nenhum monitor encontrado no arquivo {file['name']}.", flush=True)

            for monitor in monitors:
                monitors_data.append({
                    "Nome_PC": pc_name,
                    "Data_Snapshot": snapshot_date,
                    "Numero_Monitor": monitor["Numero_Monitor"],
                    "Modelo": monitor["Modelo"],
                    "Serial": monitor["Serial"]
                })
        except Exception as e:
            print(f"[AVISO] Erro ao processar arquivo {file['name']}: {e}", flush=True)
            continue

    if not monitors_data:
        print("[AVISO] Nenhum registro de monitores lido dos snapshots de hardware.", flush=True)
        show_windows_toast("Aviso no Snapshot", "Nenhum registro de monitores foi encontrado nos snapshots.")
    else:
        print(f"[OK] {len(monitors_data)} registros de monitores lidos dos snapshots de hardware", flush=True)

    return monitors_data


def _create_inventory_spreadsheet_with_monitors():
    """Cria/atualiza a planilha de inventario GB com a aba 'Perifericos - Monitores'"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        monitors_data = _read_monitors_from_hardware_snapshots()

        if not monitors_data:
            print("[AVISO] Nenhum dado de monitores encontrado. Planilha nao sera atualizada.", flush=True)
            return False

        script_dir = os.environ.get("SCRIPT_DIR", getattr(mod_config, "SCRIPT_DIR", r"C:\Scripts"))
        spreadsheet_path = os.path.join(script_dir, "CPFANI_Inventario_GB.xlsx")

        if os.path.exists(spreadsheet_path):
            wb = openpyxl.load_workbook(spreadsheet_path)
        else:
            wb = openpyxl.Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        for name in list(wb.sheetnames):
            if _normalize_snapshot_text(name).strip().lower() == "perifericos - monitores":
                del wb[name]

        ws = wb.create_sheet("Perifericos - Monitores")

        headers = ["Nome do PC", "Data do Snapshot", "Numero do Monitor", "Modelo", "Numero de Serie"]
        ws.append(headers)

        header_fill = PatternFill(start_color="3a86ff", end_color="3a86ff", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for item in monitors_data:
            ws.append([
                item["Nome_PC"],
                item["Data_Snapshot"],
                item["Numero_Monitor"],
                item["Modelo"],
                item["Serial"]
            ])

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 30
        ws.column_dimensions["E"].width = 25

        wb.save(spreadsheet_path)
        print(f"[OK] Planilha de inventario atualizada com {len(monitors_data)} monitores: {spreadsheet_path}", flush=True)
        return True
    except ImportError:
        print("[ERRO] openpyxl nao instalado. Instale com: pip install openpyxl", flush=True)
        return False
    except Exception as e:
        print(f"[ERRO] Falha ao criar planilha de inventario: {e}", flush=True)
        return False


def _parse_printers_from_hardware_snapshot(content):
    """Extrai dados de impressoras do conteudo do snapshot de hardware."""
    printers = []

    try:
        normalized = _normalize_snapshot_text(content)

        if "PERIFERICOS - IMPRESSORAS" not in normalized:
            return printers

        start_idx = normalized.find("PERIFERICOS - IMPRESSORAS")
        printer_section = normalized[start_idx:]

        block_pattern = re.compile(
            r"Impressora\s+\d+\s*:\s*\n(.*?)(?=\n\s*Impressora\s+\d+\s*:|\n={10,}|$)",
            re.DOTALL | re.IGNORECASE
        )

        blocks = block_pattern.findall(printer_section)

        for block in blocks:
            def field(pattern):
                m = re.search(pattern, block, re.IGNORECASE)
                if not m:
                    return "N/A"
                value = m.group(1).strip()
                return value if value else "N/A"

            nome = field(r"Nome\s*:\s*(.*?)\s*(?:\n|$)")
            porta = field(r"Porta\s*:\s*(.*?)\s*(?:\n|$)")
            driver = field(r"Driver\s*:\s*(.*?)\s*(?:\n|$)")
            ip = field(r"IP\s*:\s*(.*?)\s*(?:\n|$)")
            modelo_snmp = field(r"Modelo\s*\(SNMP\)\s*:\s*(.*?)\s*(?:\n|$)")
            serial_snmp = field(r"Serial\s*\(SNMP\)\s*:\s*(.*?)\s*(?:\n|$)")

            tipo = "Local"
            if ip and ip != "N/A" and re.fullmatch(r"\d{1,3}(?:\.\d{1,3}){3}", ip):
                tipo = "Rede"
            elif porta and "USB" in porta.upper():
                tipo = "USB"

            printers.append({
                "Nome": nome,
                "Porta": porta,
                "Driver": driver,
                "IP": ip,
                "Modelo_SNMP": modelo_snmp,
                "Serial_SNMP": serial_snmp,
                "Tipo": tipo
            })

        if not printers:
            print("[AVISO] Nenhuma impressora encontrada no snapshot apos parsing.", flush=True)
    except Exception as e:
        print(f"[AVISO] Erro ao parsear impressoras do snapshot: {e}", flush=True)

    return printers


def _read_printers_from_hardware_snapshots():
    """Le arquivos CPFANI_Hardware_Snapshot*.txt do Google Drive e extrai dados de impressoras."""
    printers_data = []

    service, files = _get_google_drive_service_and_snapshot_files()
    if service is None:
        return printers_data

    if not files:
        print("[AVISO] Nenhum arquivo de snapshot encontrado no Google Drive.", flush=True)
        show_windows_toast("Aviso no Snapshot", "Nenhum arquivo de snapshot encontrado no Google Drive.")
        return printers_data

    for file in files:
        try:
            from googleapiclient.http import MediaIoBaseDownload

            request = service.files().get_media(fileId=file["id"])
            fh = io.BytesIO()
            downloader = MediaIoBaseDownload(fh, request)
            done = False

            while done is False:
                status, done = downloader.next_chunk()

            content = fh.getvalue().decode("utf-8", errors="replace")
            normalized = _normalize_snapshot_text(content)

            pc_name_match = re.search(r"Nome_Computador\s*:\s*(.*?)\s*\n", normalized)
            pc_name = pc_name_match.group(1) if pc_name_match else "UNKNOWN"

            date_match = re.search(r"Gerado em:\s*(.*?)\s*\n", normalized)
            snapshot_date = date_match.group(1) if date_match else "N/A"

            printers = _parse_printers_from_hardware_snapshot(normalized)

            if not printers:
                print(f"[AVISO] Nenhuma impressora encontrada no arquivo {file['name']}.", flush=True)

            for printer in printers:
                printers_data.append({
                    "Nome_PC": pc_name,
                    "Data_Snapshot": snapshot_date,
                    "Nome_Impressora": printer["Nome"],
                    "Porta": printer["Porta"],
                    "Driver": printer["Driver"],
                    "IP": printer["IP"],
                    "Modelo_SNMP": printer["Modelo_SNMP"],
                    "Serial_SNMP": printer["Serial_SNMP"],
                    "Tipo": printer["Tipo"]
                })
        except Exception as e:
            print(f"[AVISO] Erro ao processar arquivo {file['name']}: {e}", flush=True)
            continue

    if not printers_data:
        print("[AVISO] Nenhum registro de impressoras lido dos snapshots de hardware.", flush=True)
        show_windows_toast("Aviso no Snapshot", "Nenhum registro de impressoras foi encontrado nos snapshots.")
    else:
        print(f"[OK] {len(printers_data)} registros de impressoras lidos dos snapshots de hardware", flush=True)

    return printers_data


def _create_inventory_spreadsheet_with_printers():
    """Cria/atualiza a planilha de inventario GB com a aba 'Perifericos - Impressoras'"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        printers_data = _read_printers_from_hardware_snapshots()

        if not printers_data:
            print("[AVISO] Nenhum dado de impressoras encontrado. Planilha nao sera atualizada.", flush=True)
            return False

        script_dir = os.environ.get("SCRIPT_DIR", getattr(mod_config, "SCRIPT_DIR", r"C:\Scripts"))
        spreadsheet_path = os.path.join(script_dir, "CPFANI_Inventario_GB.xlsx")

        if os.path.exists(spreadsheet_path):
            wb = openpyxl.load_workbook(spreadsheet_path)
        else:
            wb = openpyxl.Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        for name in list(wb.sheetnames):
            if _normalize_snapshot_text(name).strip().lower() == "perifericos - impressoras":
                del wb[name]

        ws = wb.create_sheet("Perifericos - Impressoras")

        headers = [
            "Nome do PC",
            "Data do Snapshot",
            "Nome Impressora",
            "Porta",
            "Driver",
            "IP",
            "Modelo (SNMP)",
            "Serial (SNMP)",
            "Tipo"
        ]
        ws.append(headers)

        header_fill = PatternFill(start_color="ff6b6b", end_color="ff6b6b", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for item in printers_data:
            ws.append([
                item["Nome_PC"],
                item["Data_Snapshot"],
                item["Nome_Impressora"],
                item["Porta"],
                item["Driver"],
                item["IP"],
                item["Modelo_SNMP"],
                item["Serial_SNMP"],
                item["Tipo"]
            ])

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 25
        ws.column_dimensions["E"].width = 30
        ws.column_dimensions["F"].width = 18
        ws.column_dimensions["G"].width = 25
        ws.column_dimensions["H"].width = 25
        ws.column_dimensions["I"].width = 12

        wb.save(spreadsheet_path)
        print(f"[OK] Planilha de inventario atualizada com {len(printers_data)} impressoras: {spreadsheet_path}", flush=True)
        return True
    except ImportError:
        print("[ERRO] openpyxl nao instalado. Instale com: pip install openpyxl", flush=True)
        return False
    except Exception as e:
        print(f"[ERRO] Falha ao criar planilha de inventario: {e}", flush=True)
        return False


class CPFani_GUI(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("Setup Automatizado CP Fani - V5.9.8")
        self.geometry("740x860")
        self.resizable(True, True)
        self.configure(fg_color="#121212")

        self.local_snapshot = "Nao informado"
        self.usuario_snapshot = "Nao informado"

        self._build_ui()
        self.protocol("WM_DELETE_WINDOW", self._on_closing)

    def _on_closing(self):
        """Tratamento seguro para fechamento da janela"""
        if messagebox.askokcancel("Sair", "Deseja realmente sair do setup?"):
            self.log("Interface fechada pelo usuario.", "INFO")
            self.destroy()

    def _build_ui(self):
        self.main_scroll = ctk.CTkScrollableFrame(self, fg_color="transparent")
        self.main_scroll.pack(fill="both", expand=True)

        header_frame = ctk.CTkFrame(self.main_scroll, fg_color="transparent")
        header_frame.pack(pady=10, fill="x")

        if HAS_PIL:
            logo_path = os.path.join(os.path.dirname(__file__), "resources", "logo_cpfani.png")
            if os.path.exists(logo_path):
                try:
                    img = Image.open(logo_path)
                    logo_img = ctk.CTkImage(img, size=(160, 50))
                    logo_label = ctk.CTkLabel(header_frame, image=logo_img, text="")
                    logo_label.pack(pady=(0, 10))
                except Exception as e:
                    self.log(f"Aviso: Falha ao carregar logo: {e}", "AVISO")

        ctk.CTkLabel(header_frame, text="SETUP AUTOMATIZADO CP FANI", font=("Segoe UI", 20, "bold"), text_color="#3a86ff").pack()
        ctk.CTkLabel(header_frame, text="v5.9.8  |  Gestao de Endpoints (Adaptacao Dinamica)", font=("Segoe UI", 11), text_color="#666666").pack()

        ui_frame = ctk.CTkFrame(self.main_scroll, fg_color="#1e1e1e", corner_radius=8)
        ui_frame.pack(padx=20, pady=5, fill="x")
        ctk.CTkLabel(ui_frame, text="1. Interface e Estetica", font=("", 12, "bold")).grid(row=0, column=0, sticky="w", padx=10, pady=5)

        self.bar_var = ctk.StringVar(value="nenhum")
        ctk.CTkRadioButton(ui_frame, text="Manter Atual", variable=self.bar_var, value="nenhum").grid(row=1, column=0, padx=20, pady=5, sticky="w")
        ctk.CTkRadioButton(ui_frame, text="Esquerda", variable=self.bar_var, value="left").grid(row=1, column=1, padx=20, pady=5, sticky="w")
        ctk.CTkRadioButton(ui_frame, text="Centro", variable=self.bar_var, value="center").grid(row=1, column=2, padx=20, pady=5, sticky="w")

        sec_frame = ctk.CTkFrame(self.main_scroll, fg_color="#1e1e1e", corner_radius=8)
        sec_frame.pack(padx=20, pady=5, fill="x")
        ctk.CTkLabel(sec_frame, text="2. Seguranca e Privacidade", font=("", 12, "bold")).pack(anchor="w", padx=10)

        self.sec_lgpd = ctk.BooleanVar(value=True)
        ctk.CTkCheckBox(sec_frame, text="Politicas de Privacidade/LGPD + Sincronizar NTP.br", variable=self.sec_lgpd).pack(anchor="w", padx=10, pady=4)

        self.sec_hello = ctk.BooleanVar(value=True)
        ctk.CTkCheckBox(sec_frame, text="Desabilitar Windows Hello, Biometria e Tela de Boas-Vindas", variable=self.sec_hello).pack(anchor="w", padx=10, pady=4)

        self.sec_firewall = ctk.BooleanVar(value=True)
        ctk.CTkCheckBox(sec_frame, text="Firewall: Restringir SMB/RPC apenas a Rede Local (Whitelist)", variable=self.sec_firewall).pack(anchor="w", padx=10, pady=4)

        self.sec_bloatware = ctk.BooleanVar(value=True)
        ctk.CTkCheckBox(sec_frame, text="Remocao Agressiva de Bloatware (AllUsers)", variable=self.sec_bloatware).pack(anchor="w", padx=10, pady=4)

        tasks_frame = ctk.CTkFrame(self.main_scroll, fg_color="#1e1e1e", corner_radius=8)
        tasks_frame.pack(padx=20, pady=5, fill="x")
        ctk.CTkLabel(tasks_frame, text="3. Automacao no Logon e Resiliencia", font=("", 12, "bold")).pack(anchor="w", padx=10)

        self.task_manutencao = ctk.BooleanVar(value=False)
        ctk.CTkCheckBox(tasks_frame, text="Agendar manutencao de rede (DHCP)", variable=self.task_manutencao).pack(anchor="w", padx=10, pady=2)

        self.task_instalar = ctk.BooleanVar(value=False)
        ctk.CTkCheckBox(tasks_frame, text="Agendar atualizador de software", variable=self.task_instalar).pack(anchor="w", padx=10, pady=2)

        self.task_reinicio = ctk.BooleanVar(value=False)
        ctk.CTkCheckBox(tasks_frame, text="Agendar Reinicio Diario automatico (21:00)", variable=self.task_reinicio).pack(anchor="w", padx=10, pady=2)

        self.task_watchdog = ctk.BooleanVar(value=True)
        ctk.CTkCheckBox(tasks_frame, text='Ativar "Self-Healing" (Auto-Cura / Vigilancia de Fundo)', variable=self.task_watchdog).pack(anchor="w", padx=10, pady=2)

        sw_frame = ctk.CTkFrame(self.main_scroll, fg_color="#1e1e1e", corner_radius=8)
        sw_frame.pack(padx=20, pady=5, fill="x")

        sw_header = ctk.CTkFrame(sw_frame, fg_color="transparent")
        sw_header.pack(fill="x", padx=10, pady=5)
        ctk.CTkLabel(sw_header, text="4. Softwares e Office", font=("", 12, "bold")).pack(side="left")

        btn_none = ctk.CTkButton(sw_header, text="Limpar Todos", font=("", 10), width=80, height=22, fg_color="#2b2b2b", hover_color="#3a3a3a", command=self.select_none_apps)
        btn_none.pack(side="right", padx=2)

        btn_all = ctk.CTkButton(sw_header, text="Selecionar Todos", font=("", 10), width=95, height=22, fg_color="#2b2b2b", hover_color="#3a3a3a", command=self.select_all_apps)
        btn_all.pack(side="right", padx=2)

        grid_frame = ctk.CTkFrame(sw_frame, fg_color="transparent")
        grid_frame.pack(fill="x", padx=10, pady=5)

        self.apps_to_install = SETTINGS.get("apps", {}).get("choco", [])
        self.app_vars = {}

        for i, app in enumerate(self.apps_to_install):
            v = ctk.BooleanVar(value=True)
            self.app_vars[app] = v
            ctk.CTkCheckBox(grid_frame, text=app.capitalize(), variable=v).grid(row=i//3, column=i%3, padx=10, pady=4, sticky="w")

        office_frame = ctk.CTkFrame(sw_frame, fg_color="transparent")
        office_frame.pack(fill="x", padx=10, pady=(10, 5))

        self.office_var = ctk.StringVar(value="nenhum")
        ctk.CTkRadioButton(office_frame, text="Nenhum Office", variable=self.office_var, value="nenhum").grid(row=0, column=0, padx=10, sticky="w")
        ctk.CTkRadioButton(office_frame, text="Office 2021", variable=self.office_var, value="office2021").grid(row=0, column=1, padx=10, sticky="w")
        ctk.CTkRadioButton(office_frame, text="OnlyOffice", variable=self.office_var, value="onlyoffice").grid(row=0, column=2, padx=10, sticky="w")

        driver_frame = ctk.CTkFrame(self.main_scroll, fg_color="#1e1e1e", corner_radius=8)
        driver_frame.pack(padx=20, pady=5, fill="x")
        ctk.CTkLabel(driver_frame, text="5. Gestao de Drivers", font=("", 12, "bold")).pack(anchor="w", padx=10)

        self.driver_var = ctk.StringVar(value="nenhum")
        ctk.CTkRadioButton(driver_frame, text="Ignorar", variable=self.driver_var, value="nenhum").pack(anchor="w", padx=10, pady=2)
        ctk.CTkRadioButton(driver_frame, text="Fabricante (Dell/HP/Lenovo)", variable=self.driver_var, value="fabricante").pack(anchor="w", padx=10, pady=2)
        ctk.CTkRadioButton(driver_frame, text="Windows Update (Forcar Instalacao)", variable=self.driver_var, value="wu").pack(anchor="w", padx=10, pady=2)

        status_frame = ctk.CTkFrame(self.main_scroll, fg_color="transparent")
        status_frame.pack(padx=20, pady=(10, 0), fill="x")

        self.status_label = ctk.CTkLabel(status_frame, text="A aguardar...", text_color="#00dd77", font=("", 12, "bold"))
        self.status_label.pack(side="left")

        self.progress_text = ctk.CTkLabel(status_frame, text="0%", font=("", 12, "bold"))
        self.progress_text.pack(side="right")

        self.progress = ctk.CTkProgressBar(self.main_scroll, mode="determinate", height=10, progress_color="#3a86ff")
        self.progress.pack(fill="x", padx=20, pady=5)
        self.progress.set(0)

        self.current_app_label = ctk.CTkLabel(self.main_scroll, text="", font=("", 11), text_color="#aaaaaa")
        self.current_app_label.pack(anchor="w", padx=20)

        self.log_area = ctk.CTkTextbox(self.main_scroll, fg_color="#0a0a0a", text_color="#00ff88", font=("Consolas", 11), height=120)
        self.log_area.pack(padx=20, pady=10, fill="both", expand=True)
        self.log_area.configure(state="disabled")

        self.btn_run = ctk.CTkButton(
            self.main_scroll,
            text="EXECUTAR DEPLOY",
            font=("", 14, "bold"),
            height=40,
            fg_color="#3a86ff",
            hover_color="#2a76ef",
            command=self.start_deploy
        )
        self.btn_run.pack(pady=(10, 5), padx=20, fill="x")

        self.btn_snapshot = ctk.CTkButton(
            self.main_scroll,
            text="GERAR APENAS SNAPSHOT",
            font=("", 13, "bold"),
            height=38,
            fg_color="#2b8a3e",
            hover_color="#1b7a2e",
            command=self.start_snapshot_only
        )
        self.btn_snapshot.pack(pady=(0, 10), padx=20, fill="x")

    def select_all_apps(self):
        for var in self.app_vars.values():
            var.set(True)
        self.log("Todos os softwares foram marcados.")

    def select_none_apps(self):
        for var in self.app_vars.values():
            var.set(False)
        self.log("Todos os softwares foram desmarcados.")

    def log(self, msg, level="INFO"):
        """Sistema de log robusto com timestamp e nivel"""
        ts = datetime.now().strftime("%H:%M:%S")
        log_msg = f"[{ts}] [{level}] {msg}"
        print(log_msg, flush=True)

        if level == "ERRO":
            formatted_msg = f"[{ts}] [ERRO] {msg}\n"
        elif level == "OK":
            formatted_msg = f"[{ts}] [OK] {msg}\n"
        elif level == "AVISO":
            formatted_msg = f"[{ts}] [AVISO] {msg}\n"
        else:
            formatted_msg = f"[{ts}] {msg}\n"

        self.after(0, self._log_safe, formatted_msg)

    def _log_safe(self, linha):
        """Insercao segura de log na interface"""
        try:
            self.log_area.configure(state="normal")
            self.log_area.insert("end", linha)
            self.log_area.see("end")
            self.log_area.configure(state="disabled")
        except Exception as e:
            print(f"[ERRO] Falha ao inserir log na UI: {e}", flush=True)

    def update_status(self, text, progress_value=None, current_app_text=None):
        """Atualiza status, progresso e aplicativo atual na interface"""
        try:
            self.status_label.configure(text=text)

            if progress_value is not None:
                progress_normalized = max(0, min(100, progress_value)) / 100
                self.progress.set(progress_normalized)
                self.progress_text.configure(text=f"{int(progress_value)}%")

            if current_app_text is not None:
                self.current_app_label.configure(text=current_app_text)

            self.update_idletasks()
        except Exception as e:
            self.log(f"Erro ao atualizar status: {e}", "ERRO")

    def _verify_file_sha256(self, file_path, expected_sha256):
        actual = _get_file_sha256(file_path)
        if not actual:
            self.log(f"Falha ao calcular hash SHA256 de {file_path}", "ERRO")
            return False

        if actual.upper() != expected_sha256.strip().upper():
            self.log(f"Hash SHA256 invalido para {os.path.basename(file_path)}. Esperado: {expected_sha256}, Obtido: {actual}", "ERRO")
            return False

        self.log(f"[OK] Hash SHA256 validado: {os.path.basename(file_path)}", "OK")
        return True

    def _github_headers(self, accept=None):
        headers = {
            "User-Agent": "Setup-CPFANI",
            "Accept": accept if accept else "application/vnd.github+json"
        }

        token = os.environ.get("GITHUB_TOKEN", "").strip()
        if token:
            headers["Authorization"] = f"Bearer {token}"

        return headers

    def _github_api_request(self, url, timeout=20):
        req = urllib.request.Request(url, headers=self._github_headers())
        try:
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                return json.loads(resp.read().decode("utf-8", errors="replace"))
        except urllib.error.HTTPError as e:
            self.log(f"Falha na API do GitHub ({e.code}): {e.reason}", "AVISO")
        except Exception as e:
            self.log(f"Erro ao consultar API do GitHub: {e}", "AVISO")

        return None

    def _download_text(self, url, timeout=30):
        headers = self._github_headers(accept="text/plain, application/octet-stream")
        req = urllib.request.Request(url, headers=headers)

        try:
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                return resp.read().decode("utf-8", errors="replace")
        except Exception as e:
            self.log(f"Falha ao baixar texto de {url}: {e}", "AVISO")
            return None

    def _download_file(self, url, dest_path, timeout=300):
        headers = self._github_headers(accept="application/octet-stream")
        req = urllib.request.Request(url, headers=headers)

        with urllib.request.urlopen(req, timeout=timeout) as resp:
            os.makedirs(os.path.dirname(dest_path), exist_ok=True)
            with open(dest_path, "wb") as out:
                shutil.copyfileobj(resp, out)

    def _extract_sha256_from_sum(self, content, target_file_name):
        if not content:
            return ""

        for line in content.splitlines():
            line = line.strip()
            if not line or line.startswith("#"):
                continue

            parts = line.split()
            if len(parts) >= 2:
                hash_value = parts[0].strip()
                file_name = parts[-1].strip().lstrip("*")

                if file_name.lower() == str(target_file_name).lower() and re.fullmatch(r"[0-9a-fA-F]{64}", hash_value):
                    return hash_value.upper()

        m = re.search(r"\b[0-9a-fA-F]{64}\b", content)
        if m:
            return m.group(0).upper()

        return ""

    def _get_latest_flameshot_release(self):
        api_url = "https://api.github.com/repos/flameshot-org/flameshot/releases/latest"
        data = self._github_api_request(api_url)

        if not data:
            return None

        tag = str(data.get("tag_name", "")).strip()
        if not tag:
            self.log("Release mais recente do Flameshot sem tag_name.", "AVISO")
            return None

        version = _version_to_list(tag)
        assets = data.get("assets", []) or []

        msi_url = ""
        msi_name = ""
        sha_url = ""

        for asset in assets:
            name = str(asset.get("name", "")).strip()
            url = str(asset.get("browser_download_url", "")).strip()
            lower_name = name.lower()

            if not name or not url:
                continue

            if lower_name.endswith(".msi") and "win64" in lower_name:
                msi_url = url
                msi_name = name
            elif lower_name.endswith(".msi.sha256sum") and "win64" in lower_name:
                sha_url = url

        if not msi_url:
            self.log("Nenhum asset MSI win64 encontrado na release mais recente do Flameshot.", "AVISO")
            return None

        if not sha_url and msi_name:
            expected_sha_name = f"{msi_name}.sha256sum"
            for asset in assets:
                name = str(asset.get("name", "")).strip()
                url = str(asset.get("browser_download_url", "")).strip()
                if name.lower() == expected_sha_name.lower() and url:
                    sha_url = url
                    break

        return {
            "tag": tag,
            "version": version,
            "msi_url": msi_url,
            "msi_name": msi_name,
            "sha_url": sha_url
        }

    def _download_with_validation(self, url, dest_path, min_size_mb=1, max_retries=3, timeout=300, expected_sha256=None):
        """Download robusto com validacao de tamanho, retry logic e hash opcional"""
        for attempt in range(1, max_retries + 1):
            try:
                self.log(f"Tentativa {attempt}/{max_retries}: Baixando {os.path.basename(dest_path)}...")
                os.makedirs(os.path.dirname(dest_path), exist_ok=True)

                start_time = time.time()
                self._download_file(url, dest_path, timeout=timeout)
                elapsed = time.time() - start_time

                file_size = os.path.getsize(dest_path)
                min_size_bytes = min_size_mb * 1024 * 1024

                if file_size < min_size_bytes:
                    self.log(f"Arquivo muito pequeno ({file_size} bytes < {min_size_bytes} bytes). Removendo...", "AVISO")
                    try:
                        os.remove(dest_path)
                    except Exception as e:
                        self.log(f"Falha ao remover arquivo corrompido: {e}", "AVISO")

                    if attempt < max_retries:
                        time.sleep(2)
                        continue
                    return False

                if expected_sha256:
                    if not self._verify_file_sha256(dest_path, expected_sha256):
                        try:
                            os.remove(dest_path)
                        except Exception as e:
                            self.log(f"Falha ao remover arquivo com hash invalido: {e}", "AVISO")

                        if attempt < max_retries:
                            time.sleep(2)
                            continue
                        return False
                else:
                    self.log("Hash SHA256 esperado nao informado. Validacao de integridade incompleta.", "AVISO")

                self.log(f"[OK] Download concluido: {file_size / (1024*1024):.2f} MB em {elapsed:.1f}s", "OK")
                return True
            except Exception as e:
                self.log(f"Falha na tentativa {attempt}: {e}", "ERRO")

                if os.path.exists(dest_path):
                    try:
                        os.remove(dest_path)
                    except Exception as e_rem:
                        self.log(f"Falha ao remover arquivo parcial: {e_rem}", "AVISO")

                if attempt < max_retries:
                    time.sleep(3)

        return False

    def install_smart_flameshot(self):
        """Instalacao inteligente do Flameshot comparando Chocolatey com GitHub latest"""
        self.log("Analisando repositorios do Flameshot (Chocolatey vs GitHub latest)...")

        choco_version = "0.0.0"

        try:
            res = subprocess.run(
                ["choco", "info", "flameshot", "--limit-output"],
                capture_output=True,
                text=True,
                timeout=15,
                creationflags=0x08000000 if sys.platform == "win32" else 0,
                encoding="utf-8",
                errors="replace"
            )

            if res.returncode == 0 and res.stdout:
                parts = res.stdout.strip().split("|")
                if len(parts) >= 2:
                    choco_version = parts[1]
                    self.log(f"Versao Chocolatey detectada: {choco_version}")
        except subprocess.TimeoutExpired:
            self.log("Timeout ao consultar Chocolatey. Continuando...", "AVISO")
        except Exception as e:
            self.log(f"Erro ao consultar Chocolatey: {e}", "AVISO")

        release = self._get_latest_flameshot_release()

        if not release:
            self.log("Nao foi possivel obter a release mais recente do GitHub. Fallback para Chocolatey.", "AVISO")
            return mod_instalar._choco_install("flameshot")

        github_tag = release["tag"]
        v_github = release["version"]
        v_choco = _version_to_list(choco_version) if choco_version != "0.0.0" else [0, 0, 0]

        self.log(f"Disponivel no Chocolatey: {choco_version}  |  Disponivel no GitHub: {github_tag}")

        if v_github >= v_choco:
            self.log(f"A versao {github_tag} do GitHub e a mais atual ou identica. Iniciando download via MSI...")

            expected_hash = ""

            if release.get("sha_url"):
                sum_text = self._download_text(release["sha_url"])
                expected_hash = self._extract_sha256_from_sum(sum_text, release["msi_name"])

                if expected_hash:
                    self.log(f"[OK] Hash SHA256 oficial obtido do arquivo .sha256sum: {expected_hash}", "OK")
                else:
                    self.log("Arquivo .sha256sum baixado, mas nao foi possivel extrair hash valido.", "AVISO")
            else:
                self.log("Asset .sha256sum nao encontrado na release. Tentando fontes locais/ambiente.", "AVISO")

            if not expected_hash:
                expected_hash = _get_expected_flameshot_sha256(github_tag)

            if not expected_hash:
                self.log("Hash SHA256 nao disponivel para o MSI. Use asset .sha256sum, GITHUB_TOKEN, CPFANI_FLAMESHOT_MSI_SHA256 ou sidecar. Pulando MSI.", "AVISO")
            else:
                temp_msi = os.path.join(r"C:\Users\Public\Downloads", release["msi_name"])

                if self._download_with_validation(release["msi_url"], temp_msi, min_size_mb=5, max_retries=3, expected_sha256=expected_hash):
                    try:
                        self.log("Executando instalacao silenciosa do MSI corporativo...")
                        install_res = subprocess.run(
                            ["msiexec", "/i", temp_msi, "/qn", "/norestart"],
                            capture_output=True,
                            timeout=120,
                            creationflags=0x08000000 if sys.platform == "win32" else 0,
                            encoding="utf-8",
                            errors="replace"
                        )

                        if install_res.returncode in [0, 3010]:
                            self.log(f"[OK] Flameshot {github_tag} instalado via GitHub MSI com sucesso.", "OK")
                            try:
                                os.remove(temp_msi)
                            except Exception as e:
                                self.log(f"Falha ao remover {temp_msi}: {e}", "AVISO")
                            return True
                        else:
                            self.log(f"MSI retornou codigo {install_res.returncode}", "AVISO")
                    except subprocess.TimeoutExpired:
                        self.log("Timeout na instalacao do MSI", "ERRO")
                    except Exception as e:
                        self.log(f"Erro na instalacao do MSI: {e}", "ERRO")

                    try:
                        if os.path.exists(temp_msi):
                            os.remove(temp_msi)
                    except Exception as e:
                        self.log(f"Falha ao remover {temp_msi}: {e}", "AVISO")
                else:
                    self.log("Download do MSI invalido ou indisponivel.", "AVISO")

            self.log("Fallback: Instalando via Chocolatey...", "AVISO")
        else:
            self.log("O pacote do Chocolatey e mais recente. Direcionando para o gerenciador...")

        return mod_instalar._choco_install("flameshot")

    def start_deploy(self):
        """Inicia o processo de deploy com confirmacao"""
        if not messagebox.askyesno("Confirmar", "Iniciar provisionamento (Modo Infiltrado + Self-Healing)?"):
            return

        self.btn_run.configure(state="disabled", text="EXECUTANDO...")
        self.btn_snapshot.configure(state="disabled")

        self.log_area.configure(state="normal")
        self.log_area.delete("1.0", "end")
        self.log_area.configure(state="disabled")

        thread = threading.Thread(target=self._safe_work, daemon=True)
        thread.start()

    def start_snapshot_only(self):
        """Inicia a geracao isolada do snapshot (sem deploy)"""
        if not messagebox.askyesno("Confirmar Snapshot", "Gerar apenas o snapshot de hardware (sem executar o deploy completo)?"):
            return

        self.btn_snapshot.configure(state="disabled", text="GERANDO SNAPSHOT...")
        self.btn_run.configure(state="disabled")

        self.log_area.configure(state="normal")
        self.log_area.delete("1.0", "end")
        self.log_area.configure(state="disabled")

        thread = threading.Thread(target=self._safe_snapshot_only_work, daemon=True)
        thread.start()

    def _safe_snapshot_only_work(self):
        """Wrapper seguro para _snapshot_only_work com captura de excecoes"""
        try:
            self._snapshot_only_work()
        except Exception as e:
            self.log(f"ERRO CRITICO NAO TRATADO: {str(e)}", "ERRO")
            self.log(f"Stack trace: {traceback.format_exc()}", "ERRO")
            self.after(0, self._finalizar_snapshot_only, [str(e)])

    def _snapshot_only_work(self):
        """Logica de geracao isolada do snapshot (sem deploy)"""
        erros = []
        start_time = time.time()

        try:
            self.log("Iniciando geracao isolada de snapshot...")
            self._coletar_dados_snapshot()

            total_tasks = 3
            completed = 0

            self.update_status("Gerando snapshot de hardware...", (completed / total_tasks) * 100, "Coletando dados...")
            try:
                self.log("Gerando snapshot de hardware (incluindo monitores, impressoras e adaptadores de rede)...")
                result = mod_config.run_snapshot_only(local=self.local_snapshot, usuario=self.usuario_snapshot)

                if result:
                    self.log(f"[OK] Snapshot gerado com sucesso: {result}", "OK")
                else:
                    self.log("Falha ao gerar snapshot", "ERRO")
                    erros.append("Snapshot")
            except Exception as e:
                self.log(f"Falha ao gerar snapshot: {e}", "ERRO")
                erros.append("Snapshot")

            completed += 1

            self.update_status("Atualizando planilha (Monitores)...", (completed / total_tasks) * 100, "Processando monitores...")
            try:
                self.log("Atualizando planilha de inventario GB com dados de monitores...")
                if _create_inventory_spreadsheet_with_monitors():
                    self.log("[OK] Planilha atualizada com aba de monitores", "OK")
                else:
                    self.log("Falha ao atualizar planilha (monitores)", "AVISO")
                    erros.append("Planilha Monitores")
            except Exception as e:
                self.log(f"Falha ao atualizar planilha (monitores): {e}", "ERRO")
                erros.append("Planilha Monitores")

            completed += 1

            self.update_status("Atualizando planilha (Impressoras)...", (completed / total_tasks) * 100, "Processando impressoras...")
            try:
                self.log("Atualizando planilha de inventario GB com dados de impressoras...")
                if _create_inventory_spreadsheet_with_printers():
                    self.log("[OK] Planilha atualizada com aba de impressoras", "OK")
                else:
                    self.log("Falha ao atualizar planilha (impressoras)", "AVISO")
                    erros.append("Planilha Impressoras")
            except Exception as e:
                self.log(f"Falha ao atualizar planilha (impressoras): {e}", "ERRO")
                erros.append("Planilha Impressoras")

            completed += 1

            elapsed_time = time.time() - start_time
            self.log(f"Snapshot concluido em {elapsed_time:.1f} segundos")
        except Exception as e:
            self.log(f"ERRO CRITICO: {str(e)}", "ERRO")
            self.log(f"Stack trace: {traceback.format_exc()}", "ERRO")
            erros.append("Critico")
        finally:
            self.after(0, self._finalizar_snapshot_only, erros)

    def _finalizar_snapshot_only(self, erros):
        """Finaliza a geracao isolada do snapshot e exibe resultados"""
        try:
            self.progress.set(1.0)
            self.progress_text.configure(text="100%")
            self.current_app_label.configure(text="")

            self.btn_snapshot.configure(state="normal", text="GERAR APENAS SNAPSHOT")
            self.btn_run.configure(state="normal")

            if erros:
                self.update_status(f"Snapshot concluido com {len(erros)} alerta(s)", 100)
                self.log(f"Snapshot concluido com {len(erros)} erro(s): {', '.join(erros)}", "AVISO")
                show_windows_toast("Aviso no Snapshot", f"Problemas com: {', '.join(erros)}.")
            else:
                self.update_status("Snapshot finalizado com sucesso!", 100)
                self.log("[OK] Snapshot concluido sem erros!", "OK")
                show_windows_toast("CP Fani - Snapshot", "Snapshot gerado com sucesso!")
        except Exception as e:
            self.log(f"Erro ao finalizar snapshot: {e}", "ERRO")

    def _safe_work(self):
        """Wrapper seguro para _work com captura de excecoes nao tratadas"""
        try:
            self._work()
        except Exception as e:
            self.log(f"ERRO CRITICO NAO TRATADO: {str(e)}", "ERRO")
            self.log(f"Stack trace: {traceback.format_exc()}", "ERRO")
            self.after(0, self._finalizar, ["Critico-NaoTratado"])

    def _coletar_dados_snapshot(self):
        """Abre janela modal para coletar local e usuario para o snapshot"""
        locais = [
            "BPCS – LOJA",
            "4830 – MATRIZ",
            "4842 – METRÓPOLE",
            "5152 – CORAÇÃO",
            "6105 – ASSAI",
            "6106 – DIREITA",
            "6110 – AROUCHE",
            "8001 – DOM JOSÉ",
            "12055 – SÃO BENTO",
            "11576 – D'AVÓ",
            "12605 – COOP",
            "12645 – LIGHT",
            "20371 – METRÔ LUZ",
            "21502 – BB_SBC",
            "23000 – OUTLET",
            "12056 – MARECHAL",
            "14120 – ARPEL SBC",
            "14353 – ARPEL SP",
            "23379 – Piraporinha"
        ]

        dialog = ctk.CTkToplevel(self)
        dialog.title("Dados do Snapshot")
        dialog.geometry("400x250")
        dialog.transient(self)
        dialog.grab_set()
        dialog.resizable(False, False)

        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - 200
        y = (dialog.winfo_screenheight() // 2) - 125
        dialog.geometry(f"+{x}+{y}")

        ctk.CTkLabel(dialog, text="Selecione o Local:", font=("", 12, "bold")).pack(pady=(20, 5))

        var_local = ctk.StringVar(value=locais[0])
        option_menu = ctk.CTkOptionMenu(dialog, values=locais, variable=var_local, width=300)
        option_menu.pack(pady=5)

        ctk.CTkLabel(dialog, text="Nome do Usuario:", font=("", 12, "bold")).pack(pady=(15, 5))

        entry_usuario = ctk.CTkEntry(dialog, width=300, placeholder_text="Digite o nome do usuario")
        entry_usuario.pack(pady=5)
        entry_usuario.focus_set()

        def confirmar():
            self.local_snapshot = var_local.get()
            self.usuario_snapshot = entry_usuario.get().strip() or "Nao informado"
            dialog.destroy()

        btn_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        btn_frame.pack(pady=20)

        ctk.CTkButton(btn_frame, text="Confirmar", command=confirmar, width=100).pack(side="left", padx=10)
        ctk.CTkButton(btn_frame, text="Cancelar", command=dialog.destroy, width=100, fg_color="#555555").pack(side="left", padx=10)

        self.wait_window(dialog)

    def _work(self):
        """Logica principal de deploy com tratamento robusto de erros"""
        erros = []
        start_time = time.time()

        try:
            self.log("Iniciando Deploy (Modo Infiltrado)...")
            self.log(f"Configuracoes carregadas: {len(SETTINGS.get('apps', {}).get('choco', []))} apps definidos")

            selected_apps = [app for app, v in self.app_vars.items() if v.get()]
            self.log(f"Aplicativos selecionados para instalacao: {len(selected_apps)}")

            self._coletar_dados_snapshot()

            total_tasks = 5
            total_tasks += len(selected_apps)

            if self.office_var.get() != "nenhum":
                total_tasks += 1

            if self.driver_var.get() != "nenhum":
                total_tasks += 1

            if self.task_watchdog.get():
                total_tasks += 1

            total_tasks += 3

            completed = 0

            self.update_status("Aplicando Interface e Branding...", (completed / total_tasks) * 100, "")
            try:
                self.log("Aplicando branding CP Fani...")
                mod_config.apply_cpfani_branding(self.bar_var.get())
                self.log("[OK] Branding aplicado com sucesso", "OK")
            except Exception as e:
                self.log(f"Falha ao aplicar branding: {e}", "ERRO")
                erros.append("Branding")

            completed += 1

            self.update_status("Aplicando Seguranca e LGPD...", (completed / total_tasks) * 100, "")
            try:
                self.log("Aplicando politicas de seguranca...")
                mod_config.apply_security_lgpd(apply_lgpd=self.sec_lgpd.get(), disable_hello=self.sec_hello.get())

                if self.sec_firewall.get():
                    self.log("Configurando regras de firewall...")
                    mod_config.apply_firewall_rules()

                if self.sec_bloatware.get():
                    self.log("Removendo bloatware...")
                    mod_config.remove_agressive_bloatware(SETTINGS.get("bloatware_remove", []))

                self.log("[OK] Seguranca aplicada com sucesso", "OK")
            except Exception as e:
                self.log(f"Falha ao aplicar seguranca: {e}", "ERRO")
                erros.append("Seguranca")

            completed += 1

            self.update_status("Verificando e removendo aplicativos legados...", (completed / total_tasks) * 100, "")
            try:
                self.log("Verificando aplicativos legados (TeamViewer, Lightshot)...")
                resultado_remocao = mod_config.check_and_remove_legacy_apps(["TeamViewer", "Lightshot"])

                for app, removido in resultado_remocao.items():
                    if removido:
                        self.log(f"[OK] {app} removido com sucesso.", "OK")
                    else:
                        self.log(f"{app} nao encontrado ou ja removido.", "INFO")
            except Exception as e:
                self.log(f"Falha ao verificar/remover apps legados: {e}", "ERRO")
                erros.append("Apps Legados")

            completed += 1

            self.update_status("Agendando Tarefas...", (completed / total_tasks) * 100, "")
            try:
                self.log("Configurando agendamentos...")

                if self.task_reinicio.get():
                    self.log("Agendando reinicio diario...")
                    mod_config.schedule_daily_reboot()

                if self.task_manutencao.get():
                    self.log("Agendando manutencao de rede...")
                    mod_config.schedule_manutencao_rede()

                if self.task_instalar.get():
                    self.log("Agendando atualizador...")
                    mod_config.schedule_instalar_tudo()

                self.log("[OK] Agendamentos configurados", "OK")
            except Exception as e:
                self.log(f"Falha ao configurar agendamentos: {e}", "ERRO")
                erros.append("Agendamentos")

            completed += 1

            if self.task_watchdog.get():
                self.update_status("Instalando Motor de Auto-Cura...", (completed / total_tasks) * 100, "Injetando Watchdog...")
                try:
                    self.log("Configurando self-healing...")
                    mod_config.setup_self_healing()
                    self.log("[OK] Self-healing ativado", "OK")
                except Exception as e:
                    self.log(f"Falha ao configurar self-healing: {e}", "ERRO")
                    erros.append("Self-Healing")

                completed += 1

            for idx, app in enumerate(selected_apps, 1):
                self.update_status(f"Instalando software ({idx}/{len(selected_apps)})", (completed / total_tasks) * 100, f"Processando {app.capitalize()}...")

                try:
                    if app == "flameshot":
                        self.log("Instalando Flameshot (smart install)...")
                        success = self.install_smart_flameshot()
                    else:
                        self.log(f"Instalando {app} via Chocolatey...")
                        success = mod_instalar._choco_install(app)

                    if success:
                        self.log(f"[OK] {app.capitalize()} instalado com sucesso", "OK")
                    else:
                        self.log(f"[ERRO] Falha ao instalar {app}", "ERRO")
                        erros.append(app)
                except Exception as e:
                    self.log(f"Erro critico ao instalar {app}: {e}", "ERRO")
                    erros.append(app)

                completed += 1

            self.update_status("Configurando arranque global...", (completed / total_tasks) * 100, "Configurando ferramentas de suporte...")
            try:
                self.log("Configurando aplicativos de startup...")
                mod_config.set_apps_to_startup_all_users()
                self.log("[OK] Startup configurado", "OK")
            except Exception as e:
                self.log(f"Falha ao configurar startup: {e}", "ERRO")
                erros.append("Startup Global")

            completed += 1

            if self.office_var.get() != "nenhum":
                self.update_status("Instalando Office...", (completed / total_tasks) * 100, f"Instalando {self.office_var.get()}")
                try:
                    self.log(f"Instalando {self.office_var.get()}...")
                    if not mod_instalar.install_office_suite(self.office_var.get()):
                        self.log(f"Falha ao instalar {self.office_var.get()}", "ERRO")
                        erros.append("Office")
                    else:
                        self.log("[OK] Office instalado", "OK")
                except Exception as e:
                    self.log(f"Erro ao instalar Office: {e}", "ERRO")
                    erros.append("Office")

                completed += 1

            if self.driver_var.get() != "nenhum":
                self.update_status("Instalando Drivers...", (completed / total_tasks) * 100, f"Modo: {self.driver_var.get()}")
                try:
                    if self.driver_var.get() == "fabricante":
                        self.log("Instalando drivers do fabricante...")
                        if not mod_instalar.install_manufacturer_drivers(SETTINGS):
                            self.log("Falha ao instalar drivers do fabricante", "ERRO")
                            erros.append("Drivers Fabricante")
                        else:
                            self.log("[OK] Drivers do fabricante instalados", "OK")
                    elif self.driver_var.get() == "wu":
                        self.log("Forcando Windows Update para drivers...")
                        if not mod_instalar.force_windows_update_drivers():
                            self.log("Falha ao forcar Windows Update", "ERRO")
                            erros.append("Windows Update")
                        else:
                            self.log("[OK] Windows Update executado", "OK")
                except Exception as e:
                    self.log(f"Erro ao instalar drivers: {e}", "ERRO")
                    erros.append("Drivers")

                completed += 1

            self.update_status("Gerando snapshot de hardware...", (completed / total_tasks) * 100, "")
            try:
                self.log("Gerando snapshot de hardware (incluindo monitores, impressoras e adaptadores de rede)...")
                mod_config.generate_full_snapshot(local=self.local_snapshot, usuario=self.usuario_snapshot)
                self.log("[OK] Snapshot de hardware gerado com dados completos", "OK")
            except Exception as e:
                self.log(f"Falha ao gerar snapshot de hardware: {e}", "ERRO")
                erros.append("Snapshot Hardware")

            completed += 1

            self.update_status("Atualizando planilha de inventario GB (Monitores)...", (completed / total_tasks) * 100, "Processando monitores...")
            try:
                self.log("Atualizando planilha de inventario GB com dados de monitores...")
                if _create_inventory_spreadsheet_with_monitors():
                    self.log("[OK] Planilha de inventario atualizada com aba de monitores", "OK")
                else:
                    self.log("Falha ao atualizar planilha de inventario (monitores)", "AVISO")
                    erros.append("Planilha Monitores")
            except Exception as e:
                self.log(f"Falha ao atualizar planilha de inventario (monitores): {e}", "ERRO")
                erros.append("Planilha Monitores")

            completed += 1

            self.update_status("Atualizando planilha de inventario GB (Impressoras)...", (completed / total_tasks) * 100, "Processando impressoras...")
            try:
                self.log("Atualizando planilha de inventario GB com dados de impressoras...")
                if _create_inventory_spreadsheet_with_printers():
                    self.log("[OK] Planilha de inventario atualizada com aba de impressoras", "OK")
                else:
                    self.log("Falha ao atualizar planilha de inventario (impressoras)", "AVISO")
                    erros.append("Planilha Impressoras")
            except Exception as e:
                self.log(f"Falha ao atualizar planilha de inventario (impressoras): {e}", "ERRO")
                erros.append("Planilha Impressoras")

            completed += 1

            elapsed_time = time.time() - start_time
            self.log(f"Deploy concluido em {elapsed_time:.1f} segundos")
        except Exception as e:
            self.log(f"ERRO CRITICO: {str(e)}", "ERRO")
            self.log(f"Stack trace: {traceback.format_exc()}", "ERRO")
            erros.append("Critico")
        finally:
            self.after(0, self._finalizar, erros)

    def _finalizar(self, erros):
        """Finaliza o deploy e exibe resultados"""
        try:
            self.progress.set(1.0)
            self.progress_text.configure(text="100%")
            self.current_app_label.configure(text="")

            self.btn_run.configure(state="normal", text="EXECUTAR DEPLOY")
            self.btn_snapshot.configure(state="normal")

            if erros:
                self.update_status(f"Concluido com {len(erros)} alerta(s)", 100)
                self.log(f"Deploy concluido com {len(erros)} erro(s): {', '.join(erros)}", "AVISO")
                show_windows_toast("Aviso no Provisionamento", f"Problemas com: {', '.join(erros)}.")
            else:
                self.update_status("Setup finalizado com sucesso!", 100)
                self.log("[OK] Deploy concluido sem erros!", "OK")
                show_windows_toast("CP Fani - Sucesso", "Provisionamento concluido com sucesso!")

            if messagebox.askyesno(
                "Reiniciar Computador",
                "O provisionamento do computador foi concluido com sucesso.\n\n"
                "Deseja reiniciar o computador agora para aplicar todas as diretivas de teclado de forma definitiva?"
            ):
                self.log("Forcando reinicio imediato do sistema operacional...", "INFO")
                subprocess.Popen(
                    ["shutdown", "/r", "/t", "5", "/f"],
                    creationflags=0x08000000 if sys.platform == "win32" else 0,
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.DEVNULL
                )
        except Exception as e:
            self.log(f"Erro ao finalizar: {e}", "ERRO")


if __name__ == "__main__":
    try:
        script_dir = os.environ.get("SCRIPT_DIR", getattr(mod_config, "SCRIPT_DIR", r"C:\Scripts"))
        Path(script_dir, "Logs").mkdir(parents=True, exist_ok=True)

        app = CPFani_GUI()
        app.mainloop()
    except Exception as e:
        msg = f"[ERRO CRITICO] Falha ao iniciar interface: {e}"
        print(msg, flush=True)
        print(traceback.format_exc(), flush=True)

        try:
            import tkinter as tk
            root = tk.Tk()
            root.withdraw()
            messagebox.showerror("Erro Critico", msg + "\n" + traceback.format_exc())
            root.destroy()
        except Exception:
            pass

        input("Pressione ENTER para sair...")
        sys.exit(1)